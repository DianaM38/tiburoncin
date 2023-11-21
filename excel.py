#Librerias
import os

import openpyxl

ruta = "./result/resultads.xlsx"

"""
CON ESTO CREAREMOS UN LIBRO DE EXCEL PARA GUARDAR TODOS LOS RESULTADOS DE LAS FUNCIONES.
EN LA PAGINA 1 VAN LOS RESULTADOS DE LA FUNCION 1, EN LA PAGINA 2 LOS RESULTADOS DE LA FUNCION 2 Y EN LA PAGINA 3 VAN LOS RESULTADOS DE LA OPCION 4.
LA ESTRUCTURA DE LAS TABLAS DE EXCEL ESTAN EN EL PDF

EL EXCEL SE IMPLEMENTARÁ EN EL ARCHIVO CAMBIOS.PY, ESPECIFICAMENTE EN LAS OPCIONES 1,2 Y 4

OPCION 1: SE USARA EXCEL DESPUES DE MOSTRAR LOS CAMBIOS RECIENTES DE LAS TASAS AL USUARIO.

OPCION 2: DESPUES DE HABER MOSTRADO DE TASAS DE CAMBIO EN LA FECHA INTRODUCIDO POR EL USUARIO, SE USARA EL EXCEL PARA GUARDAR ESAS TASAS EN EL LA PAGINA 2 DEL LIBRO.

OPCION 4: CADA VEZ QUE SE REALICE UNA CONVERSION SE GUARDARA EL RESULTADO EN EL PAGINA 3
DEL LIBRO DE EXCEL
------------------------------------------FIN----------------------------------------------
"""

"""
*Parametros
  Esta funcion recibe como parametros:
  
  #rates: La respuesta del api con la clave "rates"
  #nombre_hoja: El nombre de la hoja en la que se guardaran los datos

*Funcionamiento
  -Crea un libro nuevo de excel, pero si el libro ya existe, abre ese libro
  
  -Busca que la hoja en la que guardaremos la info esté en el libro, y de no ser asi,
  la crea o cambia el nombre de la pagina activa
  
  -Agrega los datos de la respuesta del api en la hoja activa si hay respuesta
  
  -Guarda el libro
"""
def guardarExcel(rates, nombre_hoja):
  try:
    libro = openpyxl.Workbook()
    if os.path.exists(ruta):
      libro = openpyxl.load_workbook(ruta)
      
    hoja = libro.active
    if hoja.title == "Sheet":
      hoja.title = nombre_hoja
    elif nombre_hoja in libro.sheetnames :
      hoja = libro.get_sheet_by_name(nombre_hoja)
    else:
      hoja = libro.create_sheet(nombre_hoja)
      
    if rates is not None:
      i = 3
      for moneda, tasa in rates.items():
        hoja["B2"] = "Moneda"
        hoja["C2"] = "Tasa de cambio"
        hoja.cell(row = i, column = 2 , value = moneda)
        hoja.cell(row = i , column = 3 , value = tasa)
        i += 1
      libro.save(ruta)
    print("Archivo guardado correctamente.")
  except Exception as e:
    print("Error al guardar el excel:", e)
    print('Favor de crear una carpeta llamada "result" :)')
    input()


"""
*Parametros
  Esta funcion recibe como parametros:
  
  #rates: La respuesta del api con la clave "rates"
  
  #divisa: La divisa con la que se realiza la tasa de cambios

*Funcionamiento
  -Crea el nombre de la hoja para las tasas de cambio
  
  -Utiliza la funcion guardar excel para guardar la información
  
"""
def guardarCambiosExcel(rates ,divisa):
  hoja = f"Tasa de cambio {divisa}"
  guardarExcel(rates, hoja)

"""
*Parametros
  Esta funcion recibe como parametros:
  
  #rates: La respuesta del api con la clave "rates"
  
  #divisa: la divisa con la que se realiza la tasa de cambios

  #Fecha: Fecha de la busqueda de la tasa

*Funcionamiento
  -Crea el nombre de la hoja para las tasas de cambio en la fecha especifica
  
  -Utiliza la funcion guardar excel para guardar la información
  
"""
def guardarCambiosTasas(rates ,divisa, fecha):
  hoja = f"Tasa de {divisa} en {fecha}"
  guardarExcel(rates, hoja)

"""
*Parametros
  Esta funcion recibe como parametros:

  #divisa: La divisa con la que se realiza la tasa de cambios

  #conversion: El resultado de la conversion

  #Otra: La moneda de conversion

*Funcionamiento
  -Crea un libro nuevo de excel, pero si el libro ya existe, abre ese libro

  -Busca que la hoja en la que guardaremos la info esté en el libro, y de no ser asi,
  la crea o cambia el nombre de la pagina activa

  -Agrega los datos de la respuesta del api en la hoja activa si hay respuesta y 
  a partir de la ultima conversion

  -Guarda el libro
"""
def guardarConversiones(divisa, cant, conversion, otra):
  nombre_hoja = "Conversiones"
  libro = openpyxl.Workbook()
  if os.path.exists(ruta):
    libro = openpyxl.load_workbook(ruta)

  hoja = libro.active
  if hoja.title == "Sheet":
    hoja.title = nombre_hoja
  elif nombre_hoja in libro.sheetnames :
    hoja = libro.get_sheet_by_name(nombre_hoja)
  else:
    hoja = libro.create_sheet(nombre_hoja)
  hoja["B2"] = "Moneda a convertir"
  hoja["C2"] = "Cantidad"
  hoja["E2"] = "Conversion"
  hoja["D2"] = "Moneda de conversion"
  i = 3
  while hoja.cell(row = i, column = 2).value != None:
    i += 1
  hoja.cell(row = i, column = 2 , value = divisa)
  hoja.cell(row = i , column = 3 , value = cant)
  hoja.cell(row = i , column = 5 , value = conversion)
  hoja.cell(row = i, column = 4 , value = otra)
  libro.save(ruta)
  print("Archivo guardado correctamente.")
    
  
